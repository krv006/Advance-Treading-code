import pdfplumber
import re
import os
import numpy as np
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

# todo 1 nechta pdf file bersak boladi
PDF_FILES = [
    "b.pdf",
    "a.pdf",
    "3-2 шт.pdf",
]

OUT_XLSX = "2550_штук.xlsx"
IMG_DIR = "barcodes"

os.makedirs(IMG_DIR, exist_ok=True)

def extract_sscc(text):
    m = re.search(r"SSCC:\s*(\d{10,})", text or "")
    return m.group(1) if m else None

def auto_crop_barcode(pil_img):
    img = pil_img.convert("L")
    arr = np.array(img)

    mask = arr < 170
    coords = np.column_stack(np.where(mask))

    if coords.size == 0:
        return pil_img

    y_min, x_min = coords.min(axis=0)
    y_max, x_max = coords.max(axis=0)

    pad = 10
    x_min = max(x_min - pad, 0)
    y_min = max(y_min - pad, 0)
    x_max = min(x_max + pad, arr.shape[1])
    y_max = min(y_max + pad, arr.shape[0])

    return pil_img.crop((x_min, y_min, x_max, y_max))


# Excel
wb = Workbook()
ws = wb.active
ws.title = "Data"

ws["A1"] = "FILE_NAME"
ws["B1"] = "SSCC"
ws["C1"] = "BARCODE"

ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 40

row = 2

# Har bir PDFni aylantiramiz
for pdf_file in PDF_FILES:
    if not os.path.exists(pdf_file):
        print(f"❌ Topilmadi: {pdf_file}")
        continue

    with pdfplumber.open(pdf_file) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text()
            sscc = extract_sscc(text)

            if not sscc:
                continue

            page_img = page.to_image(resolution=300).original

            w, h = page_img.size
            approx = page_img.crop((0, int(h * 0.5), w, h))

            clean_barcode = auto_crop_barcode(approx)

            img_path = os.path.join(
                IMG_DIR, f"{pdf_file.replace('.pdf','')}_{i}.png"
            )
            clean_barcode.save(img_path)

            # Excelga yozish
            ws[f"A{row}"] = pdf_file
            ws[f"B{row}"] = sscc

            xl_img = XLImage(img_path)
            xl_img.width = 260
            xl_img.height = 90

            ws.add_image(xl_img, f"C{row}")
            ws.row_dimensions[row].height = 75

            row += 1

wb.save(OUT_XLSX)

print(f"✅ Tayyor: {OUT_XLSX}")



git init
git add .
git commit -m "first commit"
git branch -M main
git remote add origin https://github.com/krv006/Advance-Treading-code.git
git push -u origin main